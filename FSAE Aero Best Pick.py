import os
import win32com.client
import openpyxl

# Define the path to the directory containing the simulation files
sim_dir = "C:/Path/to/Simulation/Directory"

# Create a list of simulation files in the directory
sim_files = [os.path.join(sim_dir, f) for f in os.listdir(sim_dir) if f.endswith('.sim')]

# Start Star CCM+ and create a new Excel workbook
starccm = win32com.client.Dispatch('star.common.StarCCM+')
wb = openpyxl.Workbook()
ws = wb.active

# Write headers for the data columns
ws.cell(row=1, column=1, value="Simulation File")
ws.cell(row=1, column=2, value="Coefficient of Drag")
ws.cell(row=1, column=3, value="Coefficient of Downforce")

# Initialize variables for finding the minimum and maximum coefficients
min_cd = float('inf')
min_cd_file = ""
max_cd = float('-inf')
max_cd_file = ""
min_cdf = float('inf')
min_cdf_file = ""
max_cdf = float('-inf')
max_cdf_file = ""

# Loop over the simulation files and extract data on drag and downforce coefficients
for i, sim_file in enumerate(sim_files):
    # Load the simulation file
    sim = starccm.simulation
    sim.loadSimulation(sim_file)

    # Get the drag and downforce coefficients from the simulation
    monitor = sim.getMonitorManager().getValuesMonitor()
    cd = monitor.getValue("forces_0|dragCoeff")
    cdf = monitor.getValue("forces_0|liftCoeff")

    # Write the simulation file name and coefficients to the Excel sheet
    row = i + 2
    ws.cell(row=row, column=1, value=sim_file)
    ws.cell(row=row, column=2, value=cd)
    ws.cell(row=row, column=3, value=cdf)

    # Check if this is the simulation with the lowest or highest coefficients so far
    if cd < min_cd:
        min_cd = cd
        min_cd_file = sim_file
    if cd > max_cd:
        max_cd = cd
        max_cd_file = sim_file
    if cdf < min_cdf:
        min_cdf = cdf
        min_cdf_file = sim_file
    if cdf > max_cdf:
        max_cdf = cdf
        max_cdf_file = sim_file

    # Close the simulation
    sim.close()

# Save the Excel file
excel_file = "C:/Path/to/YourExcelFile.xlsx"
wb.save(excel_file)

# Close Star CCM+
starccm.kill()

# Print the simulation files with the most and least coefficients
print("Simulation with lowest coefficient of drag:")
print(min_cd_file)
print("Simulation with highest coefficient of drag:")
print(max_cd_file)
print("Simulation with lowest coefficient of downforce:")
print(min_cdf_file)
print("Simulation with highest coefficient of downforce:")
print(max_cdf_file)
